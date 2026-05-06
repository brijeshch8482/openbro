"""Excel spreadsheet tool — read, edit, and open .xlsx files.

Uses openpyxl (install via openbro[office]).

Supports:
- open: launch in Excel
- read: read a sheet as a CSV-like grid (text)
- info: workbook stats (sheets, rows, cols)
- sheets: list sheet names
- get_cell: read one cell (A1 notation)
- set_cell: write one cell
- append_row: add a row at the bottom of a sheet
- find_replace: replace text across cells
- list: find .xlsx files in a folder

Save policy: edits save in-place by default. Pass `save_as=<path>` for a copy.
Cell addressing: A1 notation (e.g. "B5") or sheet=<name> argument.
"""

from __future__ import annotations

import os
import platform
import subprocess
from pathlib import Path

from openbro.tools.base import BaseTool, RiskLevel

OFFICE_DEPS_HINT = (
    "Word/Excel deps not installed. Run: pip install 'openbro[office]' "
    "(or pip install python-docx openpyxl)"
)


def _resolve(path: str) -> Path:
    p = Path(path).expanduser()
    if not p.is_absolute():
        p = Path.cwd() / p
    return p.resolve()


def _ensure_xlsx(path: Path) -> str | None:
    if not path.exists():
        return f"File not found: {path}"
    if path.suffix.lower() != ".xlsx":
        return (
            f"Only .xlsx is supported (got {path.suffix}). "
            "Convert .xls/.xlsm → .xlsx in Excel first (File → Save As)."
        )
    return None


def _coerce_value(s: str):
    """Best-effort coerce a string to int/float/bool, else leave as string."""
    if not isinstance(s, str):
        return s
    if s.lower() in ("true", "false"):
        return s.lower() == "true"
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return s


class ExcelTool(BaseTool):
    name = "excel"
    description = (
        "Read, edit, and open Microsoft Excel (.xlsx) workbooks. "
        "Use 'open' to launch in Excel, 'read' for sheet contents, "
        "'get_cell'/'set_cell'/'append_row'/'find_replace' to edit, "
        "'sheets' to list tabs, 'list' to find files."
    )
    risk = RiskLevel.MODERATE

    def schema(self) -> dict:
        return {
            "name": self.name,
            "description": self.description,
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": [
                            "open",
                            "read",
                            "info",
                            "sheets",
                            "get_cell",
                            "set_cell",
                            "append_row",
                            "find_replace",
                            "list",
                        ],
                        "description": (
                            "Action: open=launch, read=sheet grid, info=stats, "
                            "sheets=tab names, get_cell/set_cell=A1 cell ops, "
                            "append_row=add row, find_replace=text replace, "
                            "list=find .xlsx in folder."
                        ),
                    },
                    "file": {
                        "type": "string",
                        "description": "Absolute or relative path to the .xlsx file",
                    },
                    "sheet": {
                        "type": "string",
                        "description": "Sheet/tab name. Default: active sheet.",
                    },
                    "cell": {
                        "type": "string",
                        "description": "Cell in A1 notation (e.g. 'B5') for get_cell/set_cell",
                    },
                    "value": {
                        "type": "string",
                        "description": (
                            "New cell value (for set_cell). Auto-coerced to int/float/bool "
                            "where applicable; prefix with '=' for formula."
                        ),
                    },
                    "row": {
                        "type": "string",
                        "description": (
                            "Comma-separated row values for append_row (e.g. 'Brijesh,28,Mumbai')."
                        ),
                    },
                    "find": {"type": "string", "description": "Text to find (find_replace)"},
                    "replace": {
                        "type": "string",
                        "description": "Replacement text (find_replace)",
                    },
                    "save_as": {
                        "type": "string",
                        "description": "Optional output path. Default: edit saves in-place.",
                    },
                    "folder": {
                        "type": "string",
                        "description": "Folder to search (action=list). Default: current dir.",
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "Max rows to return for 'read' (default 50)",
                    },
                },
                "required": ["action"],
            },
        }

    def run(self, **kwargs) -> str:
        action = kwargs.get("action", "")

        if action == "list":
            return self._list(kwargs.get("folder", "."))
        if action == "open":
            return self._open(kwargs.get("file", ""))

        try:
            import openpyxl  # noqa: F401
        except ImportError:
            return OFFICE_DEPS_HINT

        if action == "read":
            return self._read(
                kwargs.get("file", ""),
                kwargs.get("sheet"),
                int(kwargs.get("max_rows") or 50),
            )
        if action == "info":
            return self._info(kwargs.get("file", ""))
        if action == "sheets":
            return self._sheets(kwargs.get("file", ""))
        if action == "get_cell":
            return self._get_cell(
                kwargs.get("file", ""), kwargs.get("cell", ""), kwargs.get("sheet")
            )
        if action == "set_cell":
            return self._set_cell(
                kwargs.get("file", ""),
                kwargs.get("cell", ""),
                kwargs.get("value", ""),
                kwargs.get("sheet"),
                kwargs.get("save_as"),
            )
        if action == "append_row":
            return self._append_row(
                kwargs.get("file", ""),
                kwargs.get("row", ""),
                kwargs.get("sheet"),
                kwargs.get("save_as"),
            )
        if action == "find_replace":
            return self._find_replace(
                kwargs.get("file", ""),
                kwargs.get("find", ""),
                kwargs.get("replace", ""),
                kwargs.get("sheet"),
                kwargs.get("save_as"),
            )
        return f"Unknown action: {action}"

    # ─── implementations ────────────────────────────────────

    def _list(self, folder: str) -> str:
        p = _resolve(folder)
        if not p.is_dir():
            return f"Folder not found: {p}"
        files = sorted(p.glob("**/*.xlsx"))
        if not files:
            return f"No .xlsx files in {p}"
        lines = [f"{f.relative_to(p)}  ({f.stat().st_size // 1024} KB)" for f in files[:50]]
        return "\n".join(lines) + (f"\n(+{len(files) - 50} more)" if len(files) > 50 else "")

    def _open(self, file_path: str) -> str:
        if not file_path:
            return "file is required."
        p = _resolve(file_path)
        if not p.exists():
            return f"File not found: {p}"
        try:
            if platform.system() == "Windows":
                os.startfile(str(p))  # type: ignore[attr-defined]
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(p)], check=False)
            else:
                subprocess.run(["xdg-open", str(p)], check=False)
            return f"Opened in Excel: {p}"
        except Exception as e:
            return f"Failed to open: {e}"

    def _read(self, file_path: str, sheet: str | None, max_rows: int) -> str:
        from openpyxl import load_workbook

        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        try:
            wb = load_workbook(str(p), data_only=True, read_only=True)
        except Exception as e:
            return f"Failed to open .xlsx: {e}"
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append("... (+ more rows, raise max_rows to see)")
                break
            rows.append(" | ".join("" if c is None else str(c) for c in row))
        wb.close()
        if not rows:
            return f"Sheet '{ws.title}' is empty."
        return f"[{ws.title}]\n" + "\n".join(rows)

    def _info(self, file_path: str) -> str:
        from openpyxl import load_workbook

        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        try:
            wb = load_workbook(str(p), data_only=True, read_only=True)
        except Exception as e:
            return f"Failed to open .xlsx: {e}"
        lines = [
            f"File: {p.name}",
            f"Size: {p.stat().st_size // 1024} KB",
            f"Sheets: {len(wb.sheetnames)}",
        ]
        for name in wb.sheetnames:
            ws = wb[name]
            lines.append(f"  - {name}: {ws.max_row} rows × {ws.max_column} cols")
        wb.close()
        return "\n".join(lines)

    def _sheets(self, file_path: str) -> str:
        from openpyxl import load_workbook

        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        try:
            wb = load_workbook(str(p), read_only=True)
        except Exception as e:
            return f"Failed to open .xlsx: {e}"
        names = wb.sheetnames
        wb.close()
        return "\n".join(f"- {n}" for n in names) if names else "(no sheets)"

    def _get_cell(self, file_path: str, cell: str, sheet: str | None) -> str:
        from openpyxl import load_workbook

        if not cell:
            return "'cell' is required (e.g. 'B5')."
        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        try:
            wb = load_workbook(str(p), data_only=True, read_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            value = ws[cell].value
            wb.close()
        except Exception as e:
            return f"Failed: {e}"
        return f"{cell} = {value!r}" if value is not None else f"{cell} is empty"

    def _set_cell(
        self,
        file_path: str,
        cell: str,
        value: str,
        sheet: str | None,
        save_as: str | None,
    ) -> str:
        from openpyxl import load_workbook

        if not cell:
            return "'cell' is required (e.g. 'B5')."
        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        try:
            wb = load_workbook(str(p))
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            # Preserve formula if user prefixes value with '='
            ws[cell] = value if value.startswith("=") else _coerce_value(value)
            out = _resolve(save_as) if save_as else p
            wb.save(str(out))
            wb.close()
        except PermissionError:
            return f"Permission denied saving {p}. Close file in Excel first."
        except Exception as e:
            return f"Failed: {e}"
        return f"Set {cell} = {value!r}. Saved: {out}"

    def _append_row(
        self,
        file_path: str,
        row: str,
        sheet: str | None,
        save_as: str | None,
    ) -> str:
        from openpyxl import load_workbook

        if not row:
            return "'row' is required (comma-separated values)."
        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        cells = [_coerce_value(v.strip()) for v in row.split(",")]
        try:
            wb = load_workbook(str(p))
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            ws.append(cells)
            out = _resolve(save_as) if save_as else p
            wb.save(str(out))
            wb.close()
        except PermissionError:
            return f"Permission denied saving {p}. Close file in Excel first."
        except Exception as e:
            return f"Failed: {e}"
        return f"Appended row to '{ws.title}' ({len(cells)} cells). Saved: {out}"

    def _find_replace(
        self,
        file_path: str,
        find: str,
        replace: str,
        sheet: str | None,
        save_as: str | None,
    ) -> str:
        from openpyxl import load_workbook

        if not find:
            return "'find' is required."
        p = _resolve(file_path)
        err = _ensure_xlsx(p)
        if err:
            return err
        try:
            wb = load_workbook(str(p))
        except Exception as e:
            return f"Failed to open .xlsx: {e}"

        sheets_to_scan = [wb[sheet]] if sheet and sheet in wb.sheetnames else wb.worksheets
        count = 0
        for ws in sheets_to_scan:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue
                    if find in cell.value:
                        n = cell.value.count(find)
                        cell.value = cell.value.replace(find, replace)
                        count += n

        if count == 0:
            wb.close()
            return f"'{find}' not found — no changes made."

        out = _resolve(save_as) if save_as else p
        try:
            wb.save(str(out))
        except PermissionError:
            return f"Permission denied saving {out}. Close file in Excel first."
        finally:
            wb.close()
        scope = f"sheet '{sheet}'" if sheet else "all sheets"
        return f"Replaced {count} occurrence(s) of '{find}' → '{replace}' in {scope}. Saved: {out}"
